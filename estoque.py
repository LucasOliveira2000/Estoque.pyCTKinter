from tkinter import*
from typing import Optional, Tuple, Union
import customtkinter as ctk
import openpyxl,xlrd
import pathlib
from tkinter import messagebox
from tkinter import ttk
from openpyxl import Workbook

ctk.set_appearance_mode("System")
ctk.set_default_color_theme("blue")

class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.layout_config()
        self.geometry()
        self.appearence()
        self.todo_sistema()

    def layout_config(self):
        self.title("Desenvolvido Por - Lucas Oliveira (81) 984367669 ")
        self.geometry("700x650")
        self.resizable(False, False)

    def appearence(self):
        self.lb_apm = ctk.CTkLabel(self, text="Tema", bg_color="transparent", text_color=["#000", "#fff"]).place(x=50, y=535)
        self.opt_apm = ctk.CTkOptionMenu(self, values=["Light", "Dark"], command=self.change_apm).place(x=50, y=570)

    def todo_sistema(self):
        frame = ctk.CTkFrame(self, width=900, height=50, corner_radius=0, bg_color="teal", fg_color="teal")
        frame.place(x=0, y=10)
        title = ctk.CTkLabel(frame, text="Sistema de Estoque", font=("Century Gothic bold",24),text_color="#fff").place(x=240, y=10)
        span = ctk.CTkLabel(self, text="Por favor, preencha todos os campos do formulário!", font=("Century Gothic bold", 16), text_color=["#000", "#fff"]).place(x=50, y=70)

        ficheiro = pathlib.Path("EstoqueTI.xlsx")

        if ficheiro.exists():
            pass
        else:
            ficheiro=Workbook()
            folha=ficheiro.active
                
            folha['A1']="Nome do Produto"
            folha['B1']="Marca do Produto"
            folha['C1']="Cor do Produto"
            folha['D1']="Chegada do Produto"
            folha['E1']="Funcionando"
            folha['F1']="Saida do Produto"
            folha['G1']="Observações"
                
            ficheiro.save("EstoqueTI.xlsx")   
        
        def submit():
           #Pegando os dados dos entrys
           name = name_value.get()
           marca = marca_value.get()
           cor = cor_value.get()
           chegada = chegada_value.get()
           gender = gender_combobox.get()
           saida = saida_value.get()
           obs = obs_entry.get(0.0, END) 

           if (name =="" or marca=="" or cor=="" or chegada==""):
            messagebox.showerror("Sistema", "ERRO!\nPor favor preencha todos os campos!") 

           ficheiro = openpyxl.load_workbook("EstoqueTI.xlsx")
           folha = ficheiro.active
           folha.cell(column=1, row=folha.max_row+1, value=name) 
           folha.cell(column=2, row=folha.max_row, value=marca) 
           folha.cell(column=3, row=folha.max_row, value=cor) 
           folha.cell(column=4, row=folha.max_row, value=chegada) 
           folha.cell(column=5, row=folha.max_row, value=gender) 
           folha.cell(column=6, row=folha.max_row, value=saida) 
           folha.cell(column=7, row=folha.max_row, value=obs)

           ficheiro.save(r"EstoqueTI.xlsx")
           messagebox.showinfo("Sistema", "Dados salvos com sucesso!")

        def clear():
          name_value.set("")
          marca_value.set("")
          cor_value.set("")
          chegada_value.set("")
          saida_value.set("")
          obs_entry.delete(0.0, END) 


        #Texts variables
        name_value = StringVar()
        marca_value = StringVar()
        cor_value = StringVar()
        chegada_value = StringVar()
        saida_value = StringVar()
        
        
         #Entrys
        name_entry = ctk.CTkEntry(self, width=350,textvariable=name_value , font=("Century Gothic bold", 16), fg_color="transparent") #width - tamanho em horizontal da caixa de mensagem
        marca_entry = ctk.CTkEntry(self, width=200,textvariable=marca_value , font=("Century Gothic bold", 16), fg_color="transparent")
        cor_entry = ctk.CTkEntry(self, width=150,textvariable=cor_value , font=("Century Gothic bold", 16), fg_color="transparent")
        chegada_entry = ctk.CTkEntry(self, width=200,textvariable=chegada_value , font=("Century Gothic bold", 16), fg_color="transparent")
        saida_entry = ctk.CTkEntry(self, width=200,textvariable=saida_value , font=("Century Gothic bold", 16), fg_color="transparent")
        #Entrada de observações
        obs_entry = ctk.CTkTextbox(self, width=500, height=150, font=("arial", 18), border_color="#aaa", border_width=2, fg_color="transparent") #height - tamanho em vertical da caixa de mensagem 
        #Combobox
        gender_combobox = ctk.CTkComboBox(self, values=["Sim", "Não"], font=("Century Gothic bold", 16), width=150)
        gender_combobox.set("Sim")

        #Labels
        lb_name = ctk.CTkLabel(self, text="Nome do Produto", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])
        lb_marca = ctk.CTkLabel(self, text="Marca do Produto", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])
        lb_cor = ctk.CTkLabel(self, text="Cor do Produto", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])
        lb_chegada = ctk.CTkLabel(self, text="Data de Chegada", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])
        lb_saida = ctk.CTkLabel(self, text="Data de Saida", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])
        lb_funcionando = ctk.CTkLabel(self, text="Funcionando?", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])
        lb_obs = ctk.CTkLabel(self, text="Observações", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])

        # Botões
        btn_salvar = ctk.CTkButton(self, text="Salvar dados".upper(), command=submit, fg_color="#151", hover_color="#131").place(x=300, y=520)
        btn_limpar = ctk.CTkButton(self, text="Limpar Campos".upper(), command=clear, fg_color="#555", hover_color="#333").place(x=500, y=520)


       #Posicionamento dos Elemenntos na  Janela
        lb_name.place(x=50, y=120)
        name_entry.place(x=50, y=150)

        lb_marca.place(x=450, y=120)
        marca_entry.place(x=450, y=150)

        lb_cor.place(x=300, y=190)
        cor_entry.place(x=300, y=220)

        lb_funcionando.place(x=500, y=190)
        gender_combobox.place(x=500, y=220)

        lb_chegada.place(x=50, y=190)
        chegada_entry.place(x=50, y=220)

        lb_saida.place(x=50, y=260) # x = 50 - 50 Fica um abaixo do outro
        saida_entry.place(x=50, y=290)

        lb_obs.place(x=50, y=350)
        obs_entry.place(x=160, y=350)  
    
    def change_apm(self, nova_aparencia):
        ctk.set_appearance_mode(nova_aparencia)



if __name__=="__main__":
    app = App()
    app.mainloop()